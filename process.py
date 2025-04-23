"""
Data processing module for extracting and transforming data from various sources.
"""
import pandas as pd
import ibis
from ibis import _, selectors as s
import wbgapi as wb
import re
from openpyxl import load_workbook
from pathlib import Path

from config import ISORA_FILES, WDI_INDICATOR_CODES, WGI_INDICATOR_LABELS, GFI_CONFIG

def snake_case(text: str) -> str:
    """Convert text to snake_case."""
    s = re.sub(r'[^a-zA-Z0-9]', '_', text)
    s = re.sub(r'([a-z0-9])([A-Z])', r'\1_\2', s)
    return re.sub(r'_+', '_', s.lower()).strip('_')

def process_isora(raw_data_path: Path) -> pd.DataFrame:
    """
    Extract and transform International Survey on Revenue Administration (ISORA) 
    data from Excel files into a standardized long format.
    
    Args:
        raw_data_path: Path to raw data directory
    
    Returns:
        pd.DataFrame: Combined DataFrame with all ISORA data in long format with columns:
                      ['country', 'year', 'value', 'indicator_code', 'indicator_label', 
                       'source', 'database', 'collection']
    """
    
    def extract_isora_sheet(
        file_name: str,
        sheet_name: str,
        header_row_indicator: int = 6,
        table_end_row: int = 172,
        country_col_idx: int = 1
    ) -> pd.DataFrame:
        """
        Extract and reshape an ISORA sheet with merged indicator/year columns into long format.

        Args:
            file_name: Name of Excel file
            sheet_name: Sheet name to process
            header_row_indicator: Row containing merged indicator labels (1-indexed)
            table_end_row: Last row of data (1-indexed)
            country_col_idx: Column index of the country column (1-indexed)

        Returns:
            pd.DataFrame: Long-format DataFrame with standardized columns
        """
        # Calculate derived row indices
        header_row_years = header_row_indicator + 1
        data_start_row = header_row_indicator + 2
        file_path = raw_data_path / file_name
        
        # Define country code assignment helper function
        def assign_country_iso3(df: pd.DataFrame) -> pd.DataFrame:
            """Map country names to ISO3 codes with fallbacks for special cases"""
            fallback_codes = {
                "Cook Islands": "COK",
                "Montserrat": "MSR", 
                "Republika Srpska": "BIH",
                "São Tomé and Príncipe": "STP",
                "Taiwan": "TWN",
                "Vietnam": "VNM"
            }
            
            return df.assign(
                country=lambda d: d["country_name"]
                    .map(lambda x: wb.economy.coder(x) if pd.notnull(x) else None)
                    .fillna(d["country_name"].map(fallback_codes))
            )

        # Step 1: Extract indicator blocks from merged cells
        workbook = load_workbook(filename=file_path, data_only=True)
        sheet = workbook[sheet_name]

        indicator_blocks = []
        for merged_range in sheet.merged_cells.ranges:
            if merged_range.min_row == header_row_indicator and merged_range.max_col > merged_range.min_col:
                label = sheet.cell(row=header_row_indicator, column=merged_range.min_col).value
                years = [
                    sheet.cell(row=header_row_years, column=col).value
                    for col in range(merged_range.min_col, merged_range.max_col + 1)
                ]
                indicator_blocks.append((label, years, merged_range.min_col, merged_range.max_col))

        # Step 2: Load the raw sheet as DataFrame (no headers)
        df = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            skiprows=data_start_row - 1,
            header=None,
            nrows=table_end_row - (data_start_row - 1),
            engine="openpyxl"
        )
        df.rename(columns={country_col_idx - 1: "country_name"}, inplace=True)

        # Step 3: Process each indicator block and combine results
        frames = []
        for label, years, min_col, max_col in indicator_blocks:
            # Convert Excel 1-based to Python 0-based indexing
            col_idxs = list(range(min_col - 1, max_col))
            col_map = {idx: year for idx, year in zip(col_idxs, years)}

            df_block = df[["country_name"] + col_idxs].rename(columns=col_map)
            melted = (
                df_block
                .melt(id_vars="country_name", var_name="year", value_name="value")
                .pipe(assign_country_iso3)
                .assign(
                    source="ISORA",
                    indicator_code=label, 
                    indicator_label=label,
                    database=file_name,
                    collection=sheet_name
                )
                .drop(columns=["country_name"])
            )
            frames.append(melted)

        return pd.concat(frames, ignore_index=True)

    # Combine data from all files and sheets
    isora_data = pd.concat([
        extract_isora_sheet(
            file_name=file_info["filepath"],
            sheet_name=sheet_name,
            header_row_indicator=sheet_config["start"],
            table_end_row=sheet_config["end"],
            country_col_idx=1
        )
        for file_info in ISORA_FILES.values()
        for sheet_name, sheet_config in file_info["sheets"].items()
    ], ignore_index=True)
    
    return isora_data

def process_wb(raw_data_path: Path) -> pd.DataFrame:
    """
    Process World Bank data sources into standardized format.
    
    Args:
        raw_data_path: Path to raw data directory
    
    Returns:
        Dataframe with combined World Bank sources
    """
    # WB PATHS
    pefa_path = raw_data_path / "WB-PEFA.xlsx"
    wbtax_path = raw_data_path / "WB_TAX CPACITY AND GAP.csv"
    wdi_path = raw_data_path / "WDI_CSV/WDICSV.csv"
    wgi_path = raw_data_path / "wgidataset.xlsx"

    # WB Pefa file
    wb_meta = pd.read_excel(pefa_path, sheet_name=1)

    pefa = (
        pd.read_excel(pefa_path, sheet_name=0)
        .drop(columns=['Attribute 1', 'Attribute 2', 'Attribute 3', 'Partner'])
        .melt(
            id_vars=['Economy ISO3', 'Economy Name', 'Indicator ID', 'Indicator'],
            value_vars=[str(year) for year in range(2005, 2021 + 1)],
            var_name='year',
            value_name='value')
        .merge(wb_meta, on='Indicator ID', how='left')
        .rename(columns=lambda col: snake_case(col))
        .rename(columns={
            'economy_iso3': 'country',  # Map country identifiers
            'indicator_id': 'indicator_code',  # Map indicator codes
            'indicator_name': 'indicator_label'  # Use indicator_name as the label
        })
        .assign(
            source="World Bank",
            database='WB-PEFA.xlsx',  # Add the database source
            collection='PEFA'  # Add the collection name
        )
        .filter([
            'source', 'database', 'collection', 'indicator_code', 'indicator_label', 
            'country', 'year', 'value'
        ])  # Select only the columns in isora schema
    )

    # WB Tax & Gap file
    taxwb = (
        pd.read_csv(wbtax_path, low_memory=False)
        .melt(
            id_vars=["Year", "indicator name", "indicator unit", "indicator code", "iso3_code"],
            value_vars=["value", "Buoyancy", "Capacity", "Gap", "Tax Revenue Percent"],
            var_name="vartype",
            value_name="value_true"
        )
        .assign(
            value=lambda d: pd.to_numeric(d["value_true"], errors="coerce"),
            source="World Bank",
            database='WB_TAX CPACITY AND GAP.xlsx',
            collection='TAXGAP',
            indicator_label=lambda d: d[["indicator name", "indicator unit", "vartype"]]
            .agg(lambda row: row.str.cat(sep=" - "), axis=1)
        )
        .drop(columns=["indicator unit", "indicator name", "vartype", "value_true"])
        .rename(columns=lambda col: snake_case(col))
        .rename(columns={'iso3_code': 'country'})
    )

    #  WB WDI file dump - uses ibis due to file size, pandas cannot handle
    wdi = (
        ibis.read_csv(wdi_path)
        .rename("snake_case")
        .filter(_.indicator_code.isin(WDI_INDICATOR_CODES))
        .pivot_longer(s.startswith(("1", "2")), names_to="year", values_to="value")
        .filter(_.value.notnull())
        .drop("country_name")
        .rename(
            country="country_code",
            indicator_label="indicator_name",
        )
        .mutate(
            source=ibis.literal("World Bank", type=str),
            database=ibis.literal("WDI", type=str),
            collection=ibis.literal("WDI", type=str),
        )
        .to_pandas()
    )

    #  WB WGI file
    wgi = (
        pd.read_excel(wgi_path)
        .filter(['code', 'year', 'indicator', 'estimate'])
        .rename(columns={
            'code': 'country',
            'indicator': 'indicator_code',
            'estimate': 'value'
        })
        .assign(
            source="World Bank",
            database='wgidataset.xlsx',
            collection='WGI',
            indicator_label=lambda df: df['indicator_code'].map(WGI_INDICATOR_LABELS)
        )
    )

    wb_nexus = pd.concat([pefa, taxwb, wdi, wgi], ignore_index=True)

    return wb_nexus

def process_gfi(raw_data_path: Path) -> pd.DataFrame:
    """
    Process Global Financial Integrity data.
    
    Args:
        raw_data_path: Path to raw data directory
    
    Returns:
        pd.DataFrame: Processed GFI data
    """
    gfi_path = raw_data_path / 'gfi trade mispricing.xlsx'

    def process_sheet(sheet: str, meta: dict) -> pd.DataFrame:
        """Process a single GFI sheet."""
        return (
            pd.read_excel(gfi_path, sheet_name=sheet, skiprows=4, engine="openpyxl")
            .drop(columns="Unnamed: 0")
            .rename(columns={"Unnamed: 1": "country_name"})
            .melt(id_vars="country_name", var_name="year", value_name="value")
            .assign(
                country=lambda d: d["country_name"].map(lambda x: wb.economy.coder(x) if pd.notnull(x) else None),
                source="GFI",
                database="gfi trade mispricing.xlsx",
                collection=sheet,
                indicator_code=meta["indicator_code"],
                indicator_label=meta["indicator_label"]
            )
            # Misspelled 'Syrua'
            .assign(country=lambda d: d["country"].fillna("VNM"))
            # Missing country name
            .query("not country.isnull() and year != 'Average'")
            .drop(columns="country_name")
        )
    
    # Combine data from all sheets
    return pd.concat([
        process_sheet(sheet, meta) for sheet, meta in GFI_CONFIG.items()
    ], ignore_index=True)

def process_usaid(raw_data_path: Path) -> pd.DataFrame:
    """
    Process USAID tax effort and buoyancy data.
    
    Args:
        raw_data_path: Path to raw data directory
    
    Returns:
        pd.DataFrame: Processed USAID data
    """
    usaid_path = raw_data_path / 'USAID tax effort and buyancy.xlsx'
    usaid_raw = pd.read_excel(usaid_path, engine='openpyxl', sheet_name='Data')
    
    return (
        usaid_raw
        .melt(
            id_vars=["country_id", "country_name", "year"],
            value_vars=usaid_raw.columns[3:23],
            var_name="indicator_label",
            value_name="value"
        )
        .assign(
            country=lambda d: d["country_name"].map(lambda x: wb.economy.coder(x) if pd.notnull(x) else None),
            source="USAID",
            database="Collecting Taxes Database (CTD)",
            collection="Collecting Taxes Database (CTD)",
            indicator_code=lambda df: 'USAID.CTD.' + df['indicator_label'].str.extract(r'\[([^\]]+)\]', expand=False)
        )
        # World Bank: Viet Nam
        .assign(country=lambda d: d["country"].fillna("VNM"))
        .drop(columns=["country_id", "country_name"])
    )

def process_fsi(raw_data_path: Path) -> pd.DataFrame:
    """
    Process Financial Secrecy Index data.
    
    Args:
        raw_data_path: Path to raw data directory
    
    Returns:
        pd.DataFrame: Processed FSI data
    """
    fsi_path = raw_data_path / 'tjn data.csv'
    fsi_raw = pd.read_csv(fsi_path)
    
    return (
        fsi_raw
        .melt(
            id_vars=["iso3"],
            value_vars=fsi_raw.columns[1:19],
            var_name="indicator_label",
            value_name="value"
        )
        .assign(
            # Return year numbers, cast int & add 2000 if only 2 digits
            year=lambda d: (
                d['indicator_label']
                .str.extract(r'(\d{2,4})')
                .astype(int)[0]
                .where(lambda s: s > 50, lambda s: s + 2000)
            ),
            source="TJN",
            database="Financial Secrecy Index (FSI)",
            collection="Financial Secrecy Index (FSI)",
            # Return all but year numbers
            indicator_code=lambda d: d['indicator_label'].str.replace(r'(\d{2,4})', '', regex=True)
        )
        .rename(columns={"iso3": "country"})
    )

def process_unodc(raw_data_path: Path) -> pd.DataFrame:
    """
    Process UNODC drug prices and seizures data.
    
    Args:
        raw_data_path: Path to raw data directory
    
    Returns:
        pd.DataFrame: Processed UNODC data
    """
    prices_path = raw_data_path / 'unodc drug prices.xlsx'
    seizures_path = raw_data_path / 'unodc drug seizures.xlsx'

    def load_unodc(path, sheet):
        """Load UNODC Excel file."""
        return pd.read_excel(
            path,
            skiprows=1,
            sheet_name=sheet,
            engine='openpyxl'
        )
    
    return (
        load_unodc(prices_path, 'Prices in USD')
        # when Unit in grams multiply by 1000
        .assign(
            Typical_USD=lambda d: d["Typical_USD"] * 1000
            if "Grams" in d["Unit"].values
            else d["Typical_USD"]
        )
        .drop(columns=["Region", "SubRegion", "DrugGroup"])
        .merge(
            load_unodc(seizures_path, 'Export'),
            right_on=["Country", "DrugName", "Reference year"],
            left_on=["Country/Territory", "Drug", "Year"],
            how="inner"
        )
        .assign(
            Total_Sale=lambda d: d["Kilograms"] * d["Typical_USD"]
        )
        .groupby(["Country/Territory", "Reference year"], as_index=False)["Total_Sale"]
        .sum()
        .rename(columns={
            "Reference year": "year",
            "Total_Sale": "value"
        })
        .assign(
            country=lambda d: d["Country/Territory"].map(lambda x: wb.economy.coder(x) if pd.notnull(x) else None),
            source="UNODC",
            database="Drug prices & seizures",
            collection="Drug prices & seizures",
            indicator_code='UNODC.DPS.losses',
            indicator_label=(
                "Monetary losses (in USD) to drug sales. Amount of drugs seized in kilograms multiplied by the drug price in kilograms. Excludes all seizures not measured in grams or kilograms."
            )
        )
        .drop(columns="Country/Territory")
    )

def clean_nexus_data(nexus: pd.DataFrame) -> pd.DataFrame:
    """
    Clean the combined nexus DataFrame.
    
    Args:
        nexus: Raw combined DataFrame
    
    Returns:
        pd.DataFrame: Cleaned DataFrame
    """
    return (
        nexus
        .dropna(subset=["value"])
        .assign(
            year=lambda df: pd.to_numeric(df["year"]).astype("Int64"),
            # convert value to string
            value_str=lambda df: df["value"].astype(str).str.strip().str.replace(",", "", regex=False),
            # replace coded missing with numeric and null values
            value=lambda df: df["value_str"].replace(
                {k: None for k in ["D", "N/A", "..", "N/D", "-", "", "o", "n"]} | {"Yes": "1", "No": "0"}
            ),
            # Keep info on missing codes
            value_meta=lambda df: df["value_str"].map(
                lambda x: x if pd.isna(pd.to_numeric(x, errors="coerce")) else None
            )
        )
        # value numeric conversion (throws error if anything invalid)
        .assign(value=lambda df: pd.to_numeric(df["value"], errors="raise"))
        .drop(columns=["value_str"])
    )

def country_classifiers(
        df: pd.DataFrame, 
        classifiers_path: Path
    ) -> pd.DataFrame:
    """Merge country classifications with input DataFrame."""

    columns_order=[
            'country_or_area',
            'iso3',
            "source", 
            "database", 
            "collection",
            'indicator_code',
            'indicator_label',
            'year',
            'value',
            'value_meta'
        ]
    
    result = (
        df
        .merge(
            pd.read_csv(classifiers_path)
            .rename(columns=snake_case),
            left_on='country',
            right_on='iso3',
            how='left'
        )
        .drop(columns="country")
    )
    
    # Log match statistics
    matches = result['iso3'].notna().sum()
    print(f"Country match: {matches}/{len(df)} records ({matches/len(df)*100:.1f}%)")
    
    return result

